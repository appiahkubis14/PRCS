# views.py

import csv
import json
import io
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, F, Value, CharField
from django.db.models.functions import Cast
from django.utils import timezone
from django.contrib.auth import get_user_model
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.models import (
    Business, Bill, Payment, Polygon, PropertyOwner, 
    Session, PassProperty, UserModel
)

# Helper function to get model class by name
def get_model_by_name(model_name):
    """Return the model class based on the string name"""
    models_map = {
        'business': Business,
        'bill': Bill,
        'payment': Payment,
        'polygon': Polygon,
        'property_owner': PropertyOwner,
        'session': Session,
        'pass_property': PassProperty,
    }
    return models_map.get(model_name)

# Helper function to get model fields with labels
def get_model_fields_info(model_name):
    """Return field information for a given model"""
    fields_info = {
        'business': [
            {'value': 'account_number', 'label': 'Account Number', 'type': 'text'},
            {'value': 'business_name', 'label': 'Business Name', 'type': 'text'},
            {'value': 'owner_name', 'label': 'Owner Name', 'type': 'text'},
            {'value': 'email', 'label': 'Email', 'type': 'text'},
            {'value': 'phone_number', 'label': 'Phone Number', 'type': 'text'},
            {'value': 'location', 'label': 'Location', 'type': 'text'},
            {'value': 'street_name', 'label': 'Street Name', 'type': 'text'},
            {'value': 'house_number', 'label': 'House Number', 'type': 'text'},
            {'value': 'digital_address', 'label': 'Digital Address', 'type': 'text'},
            {'value': 'flat_rate', 'label': 'Flat Rate', 'type': 'number'},
            {'value': 'created_at', 'label': 'Created At', 'type': 'date'},
            {'value': 'updated_at', 'label': 'Updated At', 'type': 'date'},
        ],
        'bill': [
            {'value': 'bill_number', 'label': 'Bill Number', 'type': 'text'},
            {'value': 'billing_year', 'label': 'Billing Year', 'type': 'number'},
            {'value': 'bill_type', 'label': 'Bill Type', 'type': 'text'},
            {'value': 'owner_name', 'label': 'Owner Name', 'type': 'text'},
            {'value': 'owner_email', 'label': 'Owner Email', 'type': 'text'},
            {'value': 'amount', 'label': 'Amount', 'type': 'number'},
            {'value': 'total_due', 'label': 'Total Due', 'type': 'number'},
            {'value': 'amount_paid', 'label': 'Amount Paid', 'type': 'number'},
            {'value': 'status', 'label': 'Status', 'type': 'text'},
            {'value': 'due_date', 'label': 'Due Date', 'type': 'date'},
            {'value': 'issued_at', 'label': 'Issued At', 'type': 'date'},
        ],
        'payment': [
            {'value': 'amount', 'label': 'Amount', 'type': 'number'},
            {'value': 'method', 'label': 'Payment Method', 'type': 'text'},
            {'value': 'reference', 'label': 'Reference', 'type': 'text'},
            {'value': 'receipt_number', 'label': 'Receipt Number', 'type': 'text'},
            {'value': 'status', 'label': 'Status', 'type': 'text'},
            {'value': 'paid_at', 'label': 'Paid At', 'type': 'date'},
            {'value': 'created_at', 'label': 'Created At', 'type': 'date'},
        ],
        'polygon': [
            {'value': 'division', 'label': 'Division', 'type': 'number'},
            {'value': 'block', 'label': 'Block', 'type': 'number'},
            {'value': 'property', 'label': 'Property Number', 'type': 'number'},
            {'value': 'g_code', 'label': 'G-Code', 'type': 'text'},
            {'value': 'district', 'label': 'District', 'type': 'text'},
            {'value': 'region', 'label': 'Region', 'type': 'text'},
            {'value': 'street', 'label': 'Street', 'type': 'text'},
            {'value': 'address', 'label': 'Address', 'type': 'text'},
            {'value': 'status', 'label': 'Status', 'type': 'text'},
            {'value': 'area_in_me', 'label': 'Area (sqm)', 'type': 'number'},
        ],
        'property_owner': [
            {'value': 'title', 'label': 'Title', 'type': 'text'},
            {'value': 'owner_name', 'label': 'Owner Name', 'type': 'text'},
            {'value': 'contact_number', 'label': 'Contact Number', 'type': 'text'},
            {'value': 'email', 'label': 'Email', 'type': 'text'},
            {'value': 'ghana_card_number', 'label': 'Ghana Card Number', 'type': 'text'},
            {'value': 'location', 'label': 'Location', 'type': 'text'},
            {'value': 'property_type', 'label': 'Property Type', 'type': 'text'},
            {'value': 'occupier', 'label': 'Occupier Type', 'type': 'text'},
            {'value': 'payment_method', 'label': 'Payment Method', 'type': 'text'},
            {'value': 'is_verified', 'label': 'Verified', 'type': 'boolean'},
        ],
        'session': [
            {'value': 'status', 'label': 'Status', 'type': 'text'},
            {'value': 'submitted_at', 'label': 'Submitted At', 'type': 'date'},
            {'value': 'reviewed_at', 'label': 'Reviewed At', 'type': 'date'},
            {'value': 'location_lat', 'label': 'Location Latitude', 'type': 'number'},
            {'value': 'location_lng', 'label': 'Location Longitude', 'type': 'number'},
        ],
        'pass_property': [
            {'value': 'reason', 'label': 'Reason', 'type': 'text'},
            {'value': 'notes', 'label': 'Notes', 'type': 'text'},
            {'value': 'passed_at', 'label': 'Passed At', 'type': 'date'},
            {'value': 'created_at', 'label': 'Created At', 'type': 'date'},
        ],
    }
    return fields_info.get(model_name, [])

# Helper function to apply filters to queryset
def apply_filters(queryset, filters, model):
    """Apply dynamic filters to the queryset"""
    if not filters:
        return queryset
    
    for filter_item in filters:
        field = filter_item.get('field')
        operator = filter_item.get('operator')
        value = filter_item.get('value')
        
        if not field or not operator or not value:
            continue
        
        # Get the actual model field
        try:
            model_field = model._meta.get_field(field)
        except:
            continue
        
        # Apply filter based on operator
        if operator == 'contains':
            queryset = queryset.filter(**{f"{field}__icontains": value})
        elif operator == 'equals':
            queryset = queryset.filter(**{field: value})
        elif operator == 'starts_with':
            queryset = queryset.filter(**{f"{field}__istartswith": value})
        elif operator == 'ends_with':
            queryset = queryset.filter(**{f"{field}__iendswith": value})
        elif operator == 'gt':
            queryset = queryset.filter(**{f"{field}__gt": value})
        elif operator == 'lt':
            queryset = queryset.filter(**{f"{field}__lt": value})
        elif operator == 'gte':
            queryset = queryset.filter(**{f"{field}__gte": value})
        elif operator == 'lte':
            queryset = queryset.filter(**{f"{field}__lte": value})
    
    return queryset

# Helper function to serialize queryset to list of dicts
def serialize_queryset(queryset, fields):
    """Convert queryset to list of dictionaries with specified fields"""
    results = []
    for obj in queryset:
        row = {}
        for field in fields:
            try:
                value = getattr(obj, field)
                # Handle datetime objects
                if hasattr(value, 'isoformat'):
                    value = value.isoformat()
                # Handle Decimal objects
                elif hasattr(value, 'quantize'):
                    value = float(value)
                row[field] = value
            except AttributeError:
                row[field] = None
        results.append(row)
    return results

# ============================================================
# API Views for Report Builder
# ============================================================

@login_required
@require_http_methods(["POST"])
def report_preview(request):
    """
    Preview endpoint - returns first 20 records with selected fields
    """
    try:
        data = json.loads(request.body)
        model_name = data.get('model')
        fields = data.get('fields', [])
        filters = data.get('filters', [])
        limit = data.get('limit', 20)
        
        # Validate model
        model = get_model_by_name(model_name)
        if not model:
            return JsonResponse({
                'success': False,
                'error': f'Invalid model: {model_name}'
            }, status=400)
        
        # Validate fields
        model_fields_info = get_model_fields_info(model_name)
        valid_field_names = [f['value'] for f in model_fields_info]
        invalid_fields = [f for f in fields if f not in valid_field_names]
        if invalid_fields:
            return JsonResponse({
                'success': False,
                'error': f'Invalid fields: {", ".join(invalid_fields)}'
            }, status=400)
        
        # Build queryset
        queryset = model.objects.all()
        
        # Apply filters
        queryset = apply_filters(queryset, filters, model)
        
        # Get total count before limiting
        total_count = queryset.count()
        
        # Apply limit for preview
        if limit and limit > 0:
            queryset = queryset[:limit]
        
        # Serialize data
        data_list = serialize_queryset(queryset, fields)
        
        return JsonResponse({
            'success': True,
            'data': data_list,
            'total_count': total_count,
            'preview_count': len(data_list)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def report_export(request):
    """
    Export endpoint - returns file in requested format
    """
    try:
        data = json.loads(request.body)
        model_name = data.get('model')
        fields = data.get('fields', [])
        filters = data.get('filters', [])
        export_format = data.get('format', 'csv')
        limit = data.get('limit', 1000)
        
        # Validate model
        model = get_model_by_name(model_name)
        if not model:
            return JsonResponse({
                'success': False,
                'error': f'Invalid model: {model_name}'
            }, status=400)
        
        # Validate fields
        model_fields_info = get_model_fields_info(model_name)
        valid_field_names = [f['value'] for f in model_fields_info]
        invalid_fields = [f for f in fields if f not in valid_field_names]
        if invalid_fields:
            return JsonResponse({
                'success': False,
                'error': f'Invalid fields: {", ".join(invalid_fields)}'
            }, status=400)
        
        # Build queryset
        queryset = model.objects.all()
        
        # Apply filters
        queryset = apply_filters(queryset, filters, model)
        
        # Apply limit
        if limit and limit > 0:
            queryset = queryset[:limit]
        
        # Get field labels for headers
        field_labels = {}
        for f in model_fields_info:
            if f['value'] in fields:
                field_labels[f['value']] = f['label']
        
        # Serialize data
        data_list = serialize_queryset(queryset, fields)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"export_{model_name}_{timestamp}"
        
        # Export based on format
        if export_format == 'csv':
            return export_csv(data_list, fields, field_labels, filename)
        elif export_format == 'xlsx':
            return export_excel(data_list, fields, field_labels, filename)
        elif export_format == 'json':
            return export_json(data_list, fields, field_labels, filename)
        elif export_format == 'xml':
            return export_xml(data_list, fields, field_labels, filename, model_name)
        else:
            return JsonResponse({
                'success': False,
                'error': f'Unsupported export format: {export_format}'
            }, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def export_csv(data_list, fields, field_labels, filename):
    """Export data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = [field_labels.get(field, field) for field in fields]
    writer.writerow(headers)
    
    # Write data
    for row in data_list:
        row_data = [row.get(field, '') for field in fields]
        writer.writerow(row_data)
    
    return response


def export_excel(data_list, fields, field_labels, filename):
    """Export data as Excel (XLSX)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = [field_labels.get(field, field) for field in fields]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, row in enumerate(data_list, 2):
        for col_idx, field in enumerate(fields, 1):
            value = row.get(field, '')
            # Handle datetime objects
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    
    return response


def export_json(data_list, fields, field_labels, filename):
    """Export data as JSON"""
    # Prepare data with labels
    formatted_data = []
    for row in data_list:
        formatted_row = {}
        for field in fields:
            label = field_labels.get(field, field)
            value = row.get(field, '')
            # Handle datetime objects
            if hasattr(value, 'isoformat'):
                value = value.isoformat()
            formatted_row[label] = value
        formatted_data.append(formatted_row)
    
    response = HttpResponse(content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
    response.write(json.dumps(formatted_data, indent=2, default=str))
    
    return response


def export_xml(data_list, fields, field_labels, filename, model_name):
    """Export data as XML"""
    root = ET.Element("export")
    root.set("model", model_name)
    root.set("exported_at", datetime.now().isoformat())
    root.set("record_count", str(len(data_list)))
    
    for row in data_list:
        record = ET.SubElement(root, "record")
        for field in fields:
            field_elem = ET.SubElement(record, field)
            field_elem.set("label", field_labels.get(field, field))
            value = row.get(field, '')
            # Handle None values
            if value is None:
                field_elem.text = ""
            else:
                # Convert to string for XML
                field_elem.text = str(value)
    
    # Convert to pretty XML
    rough_string = ET.tostring(root, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml = reparsed.toprettyxml(indent="  ")
    
    response = HttpResponse(content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xml"'
    response.write(pretty_xml)
    
    return response


@login_required
@require_http_methods(["GET"])
def get_model_fields(request):
    """
    Get available fields for a model
    """
    model_name = request.GET.get('model')
    if not model_name:
        return JsonResponse({
            'success': False,
            'error': 'Model name is required'
        }, status=400)
    
    fields_info = get_model_fields_info(model_name)
    if not fields_info:
        return JsonResponse({
            'success': False,
            'error': f'Invalid model: {model_name}'
        }, status=400)
    
    return JsonResponse({
        'success': True,
        'model': model_name,
        'fields': fields_info
    })


@login_required
@require_http_methods(["GET"])
def get_export_models(request):
    """
    Get list of available models for export
    """
    models_list = [
        {'id': 'business', 'name': 'Businesses', 'description': 'Business information and details'},
        {'id': 'bill', 'name': 'Bills', 'description': 'Billing records and invoices'},
        {'id': 'payment', 'name': 'Payments', 'description': 'Payment transactions and receipts'},
        {'id': 'polygon', 'name': 'Properties', 'description': 'Property information and locations'},
        {'id': 'property_owner', 'name': 'Property Owners', 'description': 'Property owner information'},
        {'id': 'session', 'name': 'Collection Sessions', 'description': 'Data collection sessions'},
        {'id': 'pass_property', 'name': 'Passed Properties', 'description': 'Properties marked as passed'},
    ]
    
    return JsonResponse({
        'success': True,
        'models': models_list
    })


# ============================================================
# Report Builder View
# ============================================================

@login_required
def report_builder(request):
    """
    Render the report builder page
    """
    from django.shortcuts import render
    
    context = {
        'title': 'Report Builder',
        'models': [
            {'id': 'business', 'name': 'Businesses'},
            {'id': 'bill', 'name': 'Bills'},
            {'id': 'payment', 'name': 'Payments'},
            {'id': 'polygon', 'name': 'Properties'},
            {'id': 'property_owner', 'name': 'Property Owners'},
            {'id': 'session', 'name': 'Collection Sessions'},
            {'id': 'pass_property', 'name': 'Passed Properties'},
        ],
        'export_formats': [
            {'id': 'csv', 'name': 'CSV', 'icon': 'file-csv'},
            {'id': 'xlsx', 'name': 'Excel', 'icon': 'file-excel'},
            {'id': 'json', 'name': 'JSON', 'icon': 'code'},
            {'id': 'xml', 'name': 'XML', 'icon': 'file-code'},
        ]
    }
    
    return render(request, 'core/main/exports/report-builder.html', context)